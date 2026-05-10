import discord
from discord.ext import commands, tasks
import asyncio
import os
import csv
import io
from datetime import datetime, timedelta, timezone
from collections import defaultdict
import sqlite3

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==================== CONFIG ====================
DISCORD_TOKEN         = os.environ.get("DISCORD_TOKEN")
APPROVAL_CHANNELS     = ["ssrp-approval", "outrider-ssrp-approval"]
REPORT_CHANNEL_NAME   = "leaderboard-ssrp-command"
INACTIVE_CHANNEL_NAME = "inactive-permission"
SUBMIT_COOLDOWN       = 30
INACTIVE_CHECK_HOUR   = 9
INACTIVE_CHECK_WEEKDAY= 0
ADMIN_ROLE_ID         = 1358301859663839374
LOG_CHANNEL_NAME      = "ssrp-check"

# ─── EDIT TEKS PENGUMUMAN SENIN DI SINI ──────────────
ANNOUNCE_TEXT = """📢 **PENGUMUMAN SSRP MINGGUAN!**

⚠️ Poin minggu ini akan segera **direset hari Senin ini**!
Pastikan kamu sudah submit SSRP kamu sebelum reset ya!

🎭 Submit screenshot SSRP kamu di channel ini sekarang!
🏆 Cek poin kamu dengan `!point`
📊 Cek leaderboard dengan `!lb`"""
# ═════════════════════════════════════════════════════
# ================================================

intents = discord.Intents.default()
intents.message_content = True
intents.members = True

bot = commands.Bot(command_prefix="!", intents=intents)
last_submit: dict[str, datetime] = defaultdict(lambda: datetime.min.replace(tzinfo=timezone.utc))
pending_removals: dict[str, dict] = {}

# ─── Database ─────────────────────────────────────────────────────────────────
DB = "/data/ssrp_points.db"

def init_db():
    os.makedirs("/data", exist_ok=True)
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS points (
        user_id TEXT PRIMARY KEY, username TEXT,
        total_points INTEGER DEFAULT 0, week_points INTEGER DEFAULT 0,
        last_reset TEXT, last_submit TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS submissions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id TEXT, username TEXT, submitted_at TEXT,
        photo_count INTEGER, is_valid INTEGER, reason TEXT
    )""")
    conn.commit(); conn.close()

def upsert_user(user_id, username):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT user_id FROM points WHERE user_id=?", (user_id,))
    if not c.fetchone():
        c.execute("INSERT INTO points (user_id,username,total_points,week_points,last_reset,last_submit) VALUES (?,?,0,0,?,NULL)",
                  (user_id, username, datetime.now(timezone.utc).isoformat()))
    conn.commit(); conn.close()

def add_point(user_id, username):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    now = datetime.now(timezone.utc)
    c.execute("SELECT week_points, last_reset FROM points WHERE user_id=?", (user_id,))
    row = c.fetchone()
    if row:
        last_reset = datetime.fromisoformat(row[1])
        if last_reset.tzinfo is None:
            last_reset = last_reset.replace(tzinfo=timezone.utc)
        if now - last_reset > timedelta(days=7):
            c.execute("UPDATE points SET week_points=0, last_reset=? WHERE user_id=?",
                      (now.isoformat(), user_id))
    c.execute("""UPDATE points SET
        total_points=total_points+1, week_points=week_points+1,
        username=?, last_submit=? WHERE user_id=?""",
        (username, now.isoformat(), user_id))
    conn.commit()
    c.execute("SELECT total_points, week_points FROM points WHERE user_id=?", (user_id,))
    result = c.fetchone()
    conn.close()
    return result

def log_submission(user_id, username, photo_count, is_valid, reason):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("INSERT INTO submissions (user_id,username,submitted_at,photo_count,is_valid,reason) VALUES (?,?,?,?,?,?)",
              (user_id, username, datetime.now(timezone.utc).isoformat(), photo_count, int(is_valid), reason))
    conn.commit(); conn.close()

def get_all_points():
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT user_id, username, total_points, week_points, last_submit FROM points ORDER BY total_points DESC")
    rows = c.fetchall()
    conn.close()
    return rows

def get_user_point(user_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT total_points, week_points, last_submit FROM points WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    return row

def get_user_rank(user_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT user_id FROM points ORDER BY total_points DESC")
    rows = c.fetchall()
    conn.close()
    for i, (uid,) in enumerate(rows, 1):
        if uid == user_id:
            return i, len(rows)
    return None, None

def get_inactive_users(days=7):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    threshold = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    c.execute("""SELECT user_id, username, last_submit FROM points
                 WHERE last_submit IS NULL OR last_submit < ?""", (threshold,))
    rows = c.fetchall()
    conn.close()
    return rows

def manual_add_point(user_id, username, amount=1):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT user_id FROM points WHERE user_id=?", (user_id,))
    if not c.fetchone():
        c.execute("INSERT INTO points (user_id,username,total_points,week_points,last_reset,last_submit) VALUES (?,?,0,0,?,NULL)",
                  (user_id, username, datetime.now(timezone.utc).isoformat()))
    c.execute("UPDATE points SET total_points=total_points+?, week_points=week_points+?, username=? WHERE user_id=?",
              (amount, amount, username, user_id))
    conn.commit()
    c.execute("SELECT total_points, week_points FROM points WHERE user_id=?", (user_id,))
    result = c.fetchone()
    conn.close()
    return result

def manual_remove_point(user_id, amount=1):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT total_points, week_points FROM points WHERE user_id=?", (user_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return None
    new_total = max(0, row[0] - amount)
    new_week  = max(0, row[1] - amount)
    c.execute("UPDATE points SET total_points=?, week_points=? WHERE user_id=?",
              (new_total, new_week, user_id))
    conn.commit()
    conn.close()
    return (new_total, new_week)

def db_reset_user(user_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT user_id FROM points WHERE user_id=?", (user_id,))
    if not c.fetchone():
        conn.close()
        return False
    c.execute("UPDATE points SET total_points=0, week_points=0, last_reset=?, last_submit=NULL WHERE user_id=?",
              (datetime.now(timezone.utc).isoformat(), user_id))
    conn.commit(); conn.close()
    return True

def db_remove_user(user_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT user_id FROM points WHERE user_id=?", (user_id,))
    if not c.fetchone():
        conn.close()
        return False
    c.execute("DELETE FROM points WHERE user_id=?", (user_id,))
    c.execute("DELETE FROM submissions WHERE user_id=?", (user_id,))
    conn.commit(); conn.close()
    return True

def get_user_photo_count(user_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT COALESCE(SUM(photo_count), 0) FROM submissions WHERE user_id=? AND is_valid=1", (user_id,))
    result = c.fetchone()[0]
    conn.close()
    return result

def get_user_submit_count(user_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM submissions WHERE user_id=? AND is_valid=1", (user_id,))
    result = c.fetchone()[0]
    conn.close()
    return result

def format_timestamp(ts_str):
    if not ts_str:
        return "-"
    try:
        dt = datetime.fromisoformat(ts_str)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        wib = dt + timedelta(hours=7)
        return wib.strftime("%d/%m/%Y %H:%M WIB")
    except:
        return ts_str

# ─── Excel Export (styling mirip screenshot) ──────────────────────────────────
def generate_excel(rows: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SSRP Points"

    # Warna tema
    COLOR_HEADER_BG  = "1C3A1C"  # hijau gelap
    COLOR_HEADER_FG  = "FFFFFF"
    COLOR_ROW_GREEN  = "2D5A2D"  # hijau baris biasa
    COLOR_ROW_ALT    = "245024"  # hijau agak beda (alternating)
    COLOR_ROW_RED    = "8B0000"  # merah untuk poin 0
    COLOR_ROW_RED_FG = "FFFFFF"
    COLOR_TOTAL_BG   = "1C3A1C"
    COLOR_TEXT       = "FFFFFF"

    thin = Border(
        left=Side(style="thin", color="3A6B3A"),
        right=Side(style="thin", color="3A6B3A"),
        top=Side(style="thin", color="3A6B3A"),
        bottom=Side(style="thin", color="3A6B3A")
    )
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    headers    = ["No", "Username", "User ID", "Total Poin", "Poin Minggu Ini", "Total Foto", "Terakhir Submit"]
    col_widths = [5,    28,         22,         14,           18,                12,            25]

    # Header row
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.font      = Font(bold=True, color=COLOR_HEADER_FG, name="Arial", size=11)
        cell.alignment = center
        cell.border    = thin
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28

    # Data rows
    for i, (uid, uname, total, week, last_sub) in enumerate(rows, 1):
        row_num     = i + 1
        photo_count = get_user_photo_count(uid)
        last_str    = format_timestamp(last_sub)
        is_zero     = (total == 0)

        bg_color = COLOR_ROW_RED if is_zero else (COLOR_ROW_ALT if i % 2 == 0 else COLOR_ROW_GREEN)
        fg_color = COLOR_ROW_RED_FG if is_zero else COLOR_TEXT

        values = [i, uname, uid, total, week, photo_count, last_str]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg_color)
            cell.font      = Font(name="Arial", size=10, color=fg_color,
                                  bold=(is_zero))  # bold kalau merah
            cell.border    = thin
            cell.alignment = center if col in (1, 4, 5, 6) else left
        ws.row_dimensions[row_num].height = 22

    # Total row
    total_row = len(rows) + 2
    ws.row_dimensions[total_row].height = 26
    total_labels = ["TOTAL", "", "", f"=SUM(D2:D{len(rows)+1})",
                    f"=SUM(E2:E{len(rows)+1})", f"=SUM(F2:F{len(rows)+1})", ""]
    for col, val in enumerate(total_labels, 1):
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.fill      = PatternFill("solid", fgColor=COLOR_TOTAL_BG)
        cell.font      = Font(bold=True, color=COLOR_HEADER_FG, name="Arial", size=11)
        cell.alignment = center
        cell.border    = thin

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def generate_csv(rows: list) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["No", "Username", "User ID", "Total Poin", "Poin Minggu Ini", "Total Foto", "Terakhir Submit"])
    for i, (uid, uname, total, week, last_sub) in enumerate(rows, 1):
        photo_count = get_user_photo_count(uid)
        last_str    = format_timestamp(last_sub)
        writer.writerow([i, uname, uid, total, week, photo_count, last_str])
    return buf.getvalue().encode("utf-8-sig")

# ─── Leaderboard Builder ──────────────────────────────────────────────────────
def build_leaderboard_embed(rows, page=0, per_page=15):
    start   = page * per_page
    end     = start + per_page
    chunk   = rows[start:end]
    total_p = len(rows)

    RANK_ICONS = {1:"🥇", 2:"🥈", 3:"🥉"}

    now = datetime.now(timezone.utc) + timedelta(hours=7)
    embed = discord.Embed(
        title="🏆 SSRP Leaderboard",
        color=0xF1C40F,
        timestamp=datetime.now(timezone.utc)
    )
    embed.set_footer(text=f"Halaman {page+1} • Total {total_p} member • {now.strftime('%d/%m/%Y %H:%M')} WIB")

    if not chunk:
        embed.description = "_Belum ada data._"
        return embed

    desc = ""
    for i, (uid, uname, total, week, _) in enumerate(chunk, start + 1):
        rank_icon = RANK_ICONS.get(i, f"`#{i}`")
        desc += f"{rank_icon} **{uname}** — {week} poin minggu ini _(total: {total})_\n"

    embed.description = desc
    return embed

# ─── Leaderboard View (Buttons) ───────────────────────────────────────────────
class LeaderboardView(discord.ui.View):
    def __init__(self, rows, page=0):
        super().__init__(timeout=120)
        self.rows     = rows
        self.page     = page
        self.per_page = 15
        self.max_page = max(0, (len(rows) - 1) // self.per_page)
        self._update_buttons()

    def _update_buttons(self):
        self.prev_btn.disabled = (self.page == 0)
        self.next_btn.disabled = (self.page >= self.max_page)

    @discord.ui.button(label="◀ Prev", style=discord.ButtonStyle.secondary)
    async def prev_btn(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.page = max(0, self.page - 1)
        self._update_buttons()
        await interaction.response.edit_message(
            embed=build_leaderboard_embed(self.rows, self.page), view=self
        )

    @discord.ui.button(label="🔄 Refresh", style=discord.ButtonStyle.primary)
    async def refresh_btn(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.rows = get_all_points()
        self.max_page = max(0, (len(self.rows) - 1) // self.per_page)
        self._update_buttons()
        await interaction.response.edit_message(
            embed=build_leaderboard_embed(self.rows, self.page), view=self
        )

    @discord.ui.button(label="Next ▶", style=discord.ButtonStyle.secondary)
    async def next_btn(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.page = min(self.max_page, self.page + 1)
        self._update_buttons()
        await interaction.response.edit_message(
            embed=build_leaderboard_embed(self.rows, self.page), view=self
        )

    @discord.ui.button(label="📊 Saya", style=discord.ButtonStyle.success)
    async def me_btn(self, interaction: discord.Interaction, button: discord.ui.Button):
        uid  = str(interaction.user.id)
        rank, total_members = get_user_rank(uid)
        row  = get_user_point(uid)
        if not row or not rank:
            await interaction.response.send_message("❌ Kamu belum pernah submit SSRP!", ephemeral=True)
            return
        total, week, last_sub = row
        # Pindah ke halaman yang berisi rank kamu
        target_page = (rank - 1) // self.per_page
        self.page   = target_page
        self._update_buttons()
        embed = build_leaderboard_embed(self.rows, self.page)
        embed.set_author(
            name=f"📍 Kamu di posisi #{rank} dari {total_members} member",
            icon_url=interaction.user.display_avatar.url
        )
        await interaction.response.edit_message(embed=embed, view=self)

# ─── Admin Panel View (Buttons) ───────────────────────────────────────────────
class AdminPanelView(discord.ui.View):
    def __init__(self, admin_role_id):
        super().__init__(timeout=180)
        self.admin_role_id = admin_role_id

    def _is_admin(self, interaction: discord.Interaction) -> bool:
        return any(r.id == self.admin_role_id for r in interaction.user.roles)

    @discord.ui.button(label="📊 Export Excel", style=discord.ButtonStyle.success, row=0)
    async def export_excel(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not self._is_admin(interaction):
            await interaction.response.send_message("❌ Kamu tidak punya akses!", ephemeral=True); return
        await interaction.response.defer(ephemeral=True)
        rows  = get_all_points()
        f     = discord.File(io.BytesIO(generate_excel(rows)), filename="ssrp_data.xlsx")
        await interaction.followup.send(f"📊 Export Excel — {len(rows)} member", files=[f], ephemeral=True)

    @discord.ui.button(label="📄 Export CSV", style=discord.ButtonStyle.success, row=0)
    async def export_csv(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not self._is_admin(interaction):
            await interaction.response.send_message("❌ Kamu tidak punya akses!", ephemeral=True); return
        await interaction.response.defer(ephemeral=True)
        rows = get_all_points()
        f    = discord.File(io.BytesIO(generate_csv(rows)), filename="ssrp_data.csv")
        await interaction.followup.send(f"📄 Export CSV — {len(rows)} member", files=[f], ephemeral=True)

    @discord.ui.button(label="⚠️ Cek Inactive", style=discord.ButtonStyle.danger, row=0)
    async def check_inactive_btn(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not self._is_admin(interaction):
            await interaction.response.send_message("❌ Kamu tidak punya akses!", ephemeral=True); return
        inactive = get_inactive_users(days=7)
        if not inactive:
            await interaction.response.send_message("✅ Semua member aktif minggu ini!", ephemeral=True); return
        desc = ""
        for uid, uname, last_sub in inactive[:20]:
            last_str = "Belum pernah" if not last_sub else format_timestamp(last_sub)
            desc += f"• **{uname}** — {last_str}\n"
        if len(inactive) > 20:
            desc += f"\n_...+{len(inactive)-20} lainnya_"
        embed = discord.Embed(title=f"⚠️ {len(inactive)} Member Tidak Aktif (7 Hari)", description=desc, color=0xE74C3C)
        await interaction.response.send_message(embed=embed, ephemeral=True)

    @discord.ui.button(label="🔄 Reset Mingguan", style=discord.ButtonStyle.danger, row=1)
    async def reset_week_btn(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not self._is_admin(interaction):
            await interaction.response.send_message("❌ Kamu tidak punya akses!", ephemeral=True); return
        conn = sqlite3.connect(DB)
        c    = conn.cursor()
        c.execute("UPDATE points SET week_points=0, last_reset=?", (datetime.now(timezone.utc).isoformat(),))
        conn.commit(); conn.close()
        await interaction.response.send_message("✅ **Poin mingguan semua member sudah di-reset!**", ephemeral=True)

    @discord.ui.button(label="📢 Kirim Pengumuman", style=discord.ButtonStyle.primary, row=1)
    async def announce_btn(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not self._is_admin(interaction):
            await interaction.response.send_message("❌ Kamu tidak punya akses!", ephemeral=True); return
        count = 0
        for ch_name in APPROVAL_CHANNELS:
            ch = discord.utils.get(interaction.guild.text_channels, name=ch_name)
            if ch:
                try:
                    await ch.send(ANNOUNCE_TEXT)
                    count += 1
                except Exception:
                    pass
        await interaction.response.send_message(f"✅ Pengumuman dikirim ke **{count}** channel!", ephemeral=True)

    @discord.ui.button(label="🏆 Leaderboard", style=discord.ButtonStyle.secondary, row=1)
    async def lb_btn(self, interaction: discord.Interaction, button: discord.ui.Button):
        rows  = get_all_points()
        view  = LeaderboardView(rows)
        embed = build_leaderboard_embed(rows)
        await interaction.response.send_message(embed=embed, view=view)

# ─── Auto Export Task ────────────────────────────────────────────────────────
# Jam auto export dalam WIB: 00:00, 15:15, 20:00
# Konversi ke UTC (WIB - 7): 17:00, 08:15, 13:00
AUTO_EXPORT_TIMES_UTC = [(17, 0), (8, 15), (13, 0)]

@tasks.loop(minutes=1)
async def auto_export():
    now = datetime.now(timezone.utc)
    if (now.hour, now.minute) not in AUTO_EXPORT_TIMES_UTC:
        return
    rows = get_all_points()
    if not rows:
        return
    wib_now  = now + timedelta(hours=7)
    filename = f"ssrp_data_{wib_now.strftime('%d%m%Y_%H%M')}_WIB.xlsx"
    file_bytes = generate_excel(rows)

    for guild in bot.guilds:
        log_ch = discord.utils.get(guild.text_channels, name=LOG_CHANNEL_NAME)
        if not log_ch:
            continue
        embed = discord.Embed(
            title="📊 Auto Export SSRP Data",
            description=(
                f"Export otomatis pada **{wib_now.strftime('%d/%m/%Y %H:%M')} WIB**\n"
                f"Total member: **{len(rows)}**"
            ),
            color=0x2ECC71,
            timestamp=now
        )
        embed.set_footer(text="SSRP Auto Export")
        try:
            await log_ch.send(
                embed=embed,
                file=discord.File(io.BytesIO(file_bytes), filename=filename)
            )
        except Exception as e:
            print(f"⚠️ Auto export error: {e}")

# ─── Auto Announce Task ───────────────────────────────────────────────────────
@tasks.loop(hours=1)
async def auto_announce():
    now = datetime.now(timezone.utc)
    if not (now.weekday() == 6 and now.hour == 17):
        return
    for guild in bot.guilds:
        for ch_name in APPROVAL_CHANNELS:
            ch = discord.utils.get(guild.text_channels, name=ch_name)
            if ch:
                try:
                    await ch.send(ANNOUNCE_TEXT)
                except Exception:
                    pass

# ─── Inactive Check Task ──────────────────────────────────────────────────────
@tasks.loop(hours=24)
async def check_inactive():
    now = datetime.now(timezone.utc)
    if now.weekday() != INACTIVE_CHECK_WEEKDAY or now.hour != INACTIVE_CHECK_HOUR:
        return
    for guild in bot.guilds:
        inactive_channel = discord.utils.get(guild.text_channels, name=INACTIVE_CHANNEL_NAME)
        if not inactive_channel:
            continue
        inactive = get_inactive_users(days=7)
        if not inactive:
            await inactive_channel.send("✅ **Semua member sudah submit SSRP minggu ini!**")
            continue
        chunks = [inactive[i:i+20] for i in range(0, len(inactive), 20)]
        for chunk_idx, chunk in enumerate(chunks):
            desc = ""
            for uid, uname, last_sub in chunk:
                member   = guild.get_member(int(uid))
                mention  = member.mention if member else f"**{uname}**"
                last_str = "Belum pernah" if not last_sub else format_timestamp(last_sub)
                desc    += f"• {mention} — terakhir: {last_str}\n"
            embed = discord.Embed(
                title=f"⚠️ Member Tidak Aktif SSRP (7 Hari) — Part {chunk_idx+1}",
                description=desc, color=0xE74C3C, timestamp=now
            )
            embed.set_footer(text=f"Total tidak aktif: {len(inactive)} member")
            await inactive_channel.send(embed=embed)
        for uid, uname, last_sub in inactive:
            member = guild.get_member(int(uid))
            if member:
                try:
                    await member.send(
                        f"⚠️ **Hei {member.display_name}!**\n"
                        f"Kamu belum submit SSRP dalam **7 hari terakhir**.\n"
                        f"Jangan lupa kirim screenshot SSRP kamu di server ya! 🎭"
                    )
                except Exception:
                    pass

# ─── Events ───────────────────────────────────────────────────────────────────
@bot.event
async def on_ready():
    init_db()
    check_inactive.start()
    auto_announce.start()
    auto_export.start()
    try:
        await asyncio.sleep(3)  # Tunggu bot benar-benar siap
        bot.tree.clear_commands(guild=None)
        await bot.tree.sync()
        for guild in bot.guilds:
            bot.tree.copy_global_to(guild=guild)
            await bot.tree.sync(guild=guild)
            print(f"✅ Synced to guild: {guild.name}")
        print("✅ Slash commands synced!")
    except Exception as e:
        print(f"⚠️ Slash sync error: {e}")
    print(f"✅ Bot online: {bot.user}")
    print(f"📋 Monitoring: {', '.join(['#'+c for c in APPROVAL_CHANNELS])}")

@bot.event
async def on_message(message: discord.Message):
    if message.author.bot:
        return
    if message.channel.name not in APPROVAL_CHANNELS:
        await bot.process_commands(message)
        return
    images = [a for a in message.attachments if a.content_type and a.content_type.startswith("image/")]
    if not images:
        await bot.process_commands(message)
        return
    user    = message.author
    user_id = str(user.id)
    now     = datetime.now(timezone.utc)
    elapsed = (now - last_submit[user_id]).total_seconds()
    if elapsed < SUBMIT_COOLDOWN:
        remaining = int(SUBMIT_COOLDOWN - elapsed)
        await message.add_reaction("⏳")
        await message.reply(
            f"⚠️ **{user.display_name}**, tunggu **{remaining} detik** lagi ya!\n_(Submit ini tidak dihitung)_",
            delete_after=10
        )
        return
    last_submit[user_id] = now
    upsert_user(user_id, user.display_name)
    log_submission(user_id, user.display_name, len(images), True, "Auto approved")
    total_pts, week_pts = add_point(user_id, user.display_name)
    await message.add_reaction("✅")

    report_channel = discord.utils.get(message.guild.text_channels, name=REPORT_CHANNEL_NAME) or message.channel
    msg_link       = f"https://discord.com/channels/{message.guild.id}/{message.channel.id}/{message.id}"
    channel_label  = "Outrider SSRP" if message.channel.name == "outrider-ssrp-approval" else "SSRP"
    rank, total_m  = get_user_rank(user_id)

    embed = discord.Embed(title=f"🎭 Laporan {channel_label} Baru!", color=0x57F287, timestamp=now)
    embed.set_author(name=user.display_name, icon_url=user.display_avatar.url)
    embed.add_field(name="👤 Player",           value=user.mention,                    inline=True)
    embed.add_field(name="🖼️ Jumlah Foto",      value=f"{len(images)} foto",           inline=True)
    embed.add_field(name="📅 Waktu",            value=f"<t:{int(now.timestamp())}:F>", inline=True)
    embed.add_field(name="📈 Point Minggu Ini", value=f"**{week_pts} Point**",         inline=True)
    embed.add_field(name="🏆 Total Point",      value=f"**{total_pts} Point**",        inline=True)
    embed.add_field(name="🏅 Ranking",          value=f"**#{rank}** dari {total_m}",   inline=True)
    embed.add_field(name="🔗 Link Bukti",       value=f"[Klik di sini]({msg_link})",   inline=False)
    embed.set_image(url=images[0].url)
    embed.set_footer(text=f"SSRP Checker Bot • #{message.channel.name}")
    await report_channel.send(embed=embed)
    await bot.process_commands(message)

# ─── Slash Commands ───────────────────────────────────────────────────────────
@bot.tree.command(name="announce", description="Kirim pengumuman SSRP manual ke semua channel submit")
@discord.app_commands.checks.has_any_role(ADMIN_ROLE_ID)
async def slash_announce(interaction: discord.Interaction):
    count = 0
    for ch_name in APPROVAL_CHANNELS:
        ch = discord.utils.get(interaction.guild.text_channels, name=ch_name)
        if ch:
            try:
                await ch.send(ANNOUNCE_TEXT)
                count += 1
            except Exception:
                pass
    await interaction.response.send_message(
        f"✅ Pengumuman berhasil dikirim ke **{count}** channel!", ephemeral=True
    )

@bot.tree.command(name="admin", description="Buka panel admin SSRP")
@discord.app_commands.checks.has_any_role(ADMIN_ROLE_ID)
async def slash_admin(interaction: discord.Interaction):
    embed = discord.Embed(
        title="⚙️ Admin Panel SSRP",
        description="Gunakan tombol di bawah untuk mengelola data SSRP.",
        color=0x5865F2,
        timestamp=datetime.now(timezone.utc)
    )
    embed.set_footer(text=f"Diakses oleh {interaction.user.display_name}")
    view = AdminPanelView(ADMIN_ROLE_ID)
    await interaction.response.send_message(embed=embed, view=view, ephemeral=True)

# ─── Prefix Commands ──────────────────────────────────────────────────────────
@bot.command(name="point")
async def check_point(ctx, member: discord.Member = None):
    target = member or ctx.author
    row    = get_user_point(str(target.id))
    if not row:
        await ctx.reply(f"**{target.display_name}** belum pernah submit SSRP."); return
    total, week, last_sub = row
    rank, total_m = get_user_rank(str(target.id))
    photos  = get_user_photo_count(str(target.id))
    submits = get_user_submit_count(str(target.id))
    last_str = "-"
    if last_sub:
        try: last_str = f"<t:{int(datetime.fromisoformat(last_sub).timestamp())}:R>"
        except: last_str = last_sub

    embed = discord.Embed(title="📊 SSRP Points", color=0x5865F2)
    embed.set_author(name=target.display_name, icon_url=target.display_avatar.url)
    embed.set_thumbnail(url=target.display_avatar.url)
    embed.add_field(name="🏅 Ranking",        value=f"**#{rank}** dari {total_m}", inline=True)
    embed.add_field(name="🏆 Total Poin",      value=f"**{total}**",                inline=True)
    embed.add_field(name="📈 Poin Minggu Ini", value=f"**{week}**",                 inline=True)
    embed.add_field(name="🖼️ Total Foto",      value=f"**{photos}** foto",          inline=True)
    embed.add_field(name="📋 Total Submit",    value=f"**{submits}** kali",         inline=True)
    embed.add_field(name="🕐 Terakhir Submit", value=last_str,                      inline=True)
    await ctx.reply(embed=embed)

@bot.command(name="profile")
async def profile(ctx, member: discord.Member = None):
    """Kartu profil SSRP lengkap."""
    target = member or ctx.author
    row    = get_user_point(str(target.id))
    if not row:
        await ctx.reply(f"**{target.display_name}** belum pernah submit SSRP."); return
    total, week, last_sub = row
    rank, total_m = get_user_rank(str(target.id))
    photos  = get_user_photo_count(str(target.id))
    submits = get_user_submit_count(str(target.id))

    status   = "🟢 Aktif minggu ini" if week > 0 else "🔴 Belum submit minggu ini"
    last_str = "-"
    if last_sub:
        try: last_str = f"<t:{int(datetime.fromisoformat(last_sub).timestamp())}:F>"
        except: last_str = last_sub

    embed = discord.Embed(
        title=f"🎭 Profil SSRP — {target.display_name}",
        color=0x57F287,
        timestamp=datetime.now(timezone.utc)
    )
    embed.set_thumbnail(url=target.display_avatar.url)
    embed.add_field(name="🏅 Ranking",        value=f"**#{rank}** dari {total_m}", inline=True)
    embed.add_field(name="🏆 Total Poin",      value=f"**{total}** poin",           inline=True)
    embed.add_field(name="📈 Poin Minggu Ini", value=f"**{week}** poin",            inline=True)
    embed.add_field(name="🖼️ Total Foto",      value=f"**{photos}** foto",          inline=True)
    embed.add_field(name="📋 Total Submit",    value=f"**{submits}** kali",         inline=True)
    embed.add_field(name="🔵 Status",          value=status,                        inline=True)
    embed.add_field(name="🕐 Terakhir Submit", value=last_str,                      inline=False)
    embed.set_footer(text="SSRP Checker Bot")
    await ctx.reply(embed=embed)

@bot.command(name="leaderboard", aliases=["lb"])
async def leaderboard(ctx):
    rows  = get_all_points()
    if not rows:
        await ctx.reply("Belum ada data SSRP."); return
    view  = LeaderboardView(rows)
    embed = build_leaderboard_embed(rows)
    await ctx.reply(embed=embed, view=view)

@bot.command(name="export")
@commands.has_any_role(ADMIN_ROLE_ID)
async def export_data(ctx, fmt: str = "excel"):
    rows = get_all_points()
    if not rows:
        await ctx.reply("Belum ada data untuk di-export."); return
    fmt   = fmt.lower()
    files = []
    if fmt in ("excel", "xlsx", "all"):
        files.append(discord.File(io.BytesIO(generate_excel(rows)), filename="ssrp_data.xlsx"))
    if fmt in ("csv", "all"):
        files.append(discord.File(io.BytesIO(generate_csv(rows)), filename="ssrp_data.csv"))
    if not files:
        await ctx.reply("Format tidak valid. Gunakan: `excel`, `csv`, atau `all`"); return
    await ctx.reply(f"📊 **Export SSRP Data** — {len(rows)} member", files=files)

@bot.command(name="inactive")
@commands.has_any_role(ADMIN_ROLE_ID)
async def check_inactive_cmd(ctx, days: int = 7):
    inactive = get_inactive_users(days=days)
    if not inactive:
        await ctx.reply(f"✅ Semua member sudah submit SSRP dalam {days} hari terakhir!"); return
    desc = ""
    for uid, uname, last_sub in inactive[:30]:
        member   = ctx.guild.get_member(int(uid))
        mention  = member.mention if member else f"**{uname}**"
        last_str = "Belum pernah" if not last_sub else format_timestamp(last_sub)
        desc    += f"• {mention} — terakhir: {last_str}\n"
    if len(inactive) > 30:
        desc += f"\n_...dan {len(inactive)-30} lainnya. Gunakan `!export` untuk list lengkap._"
    embed = discord.Embed(title=f"⚠️ Tidak Aktif SSRP ({days} Hari Terakhir)", description=desc, color=0xE74C3C)
    embed.set_footer(text=f"Total: {len(inactive)} member tidak aktif")
    await ctx.reply(embed=embed)

@bot.command(name="resetweek")
@commands.has_any_role(ADMIN_ROLE_ID)
async def reset_week(ctx):
    conn = sqlite3.connect(DB)
    c    = conn.cursor()
    c.execute("UPDATE points SET week_points=0, last_reset=?", (datetime.now(timezone.utc).isoformat(),))
    conn.commit(); conn.close()
    await ctx.reply("✅ **Poin mingguan semua member sudah di-reset!**")

@bot.command(name="resetuser")
@commands.has_any_role(ADMIN_ROLE_ID)
async def reset_user_cmd(ctx, member: discord.Member = None):
    if not member:
        await ctx.reply("❌ Gunakan: `!resetuser @user`"); return
    success = db_reset_user(str(member.id))
    if not success:
        await ctx.reply(f"❌ **{member.display_name}** tidak ada di database."); return
    embed = discord.Embed(title="🔄 User Di-reset!", color=0xE67E22)
    embed.set_author(name=member.display_name, icon_url=member.display_avatar.url)
    embed.add_field(name="👤 Player", value=member.mention,          inline=True)
    embed.add_field(name="🔄 Status", value="Semua poin jadi **0**", inline=True)
    embed.set_footer(text=f"Di-reset oleh {ctx.author.display_name}")
    await ctx.reply(embed=embed)
    log_channel = discord.utils.get(ctx.guild.text_channels, name=LOG_CHANNEL_NAME)
    if log_channel:
        log_embed = discord.Embed(title="🔄 Reset User — Log Admin", color=0xE67E22, timestamp=datetime.now(timezone.utc))
        log_embed.add_field(name="👮 Admin",  value=ctx.author.mention,       inline=True)
        log_embed.add_field(name="👤 Target", value=member.mention,           inline=True)
        log_embed.add_field(name="🔄 Aksi",   value="Semua poin direset → 0", inline=False)
        log_embed.set_footer(text="SSRP Admin Log")
        await log_channel.send(embed=log_embed)

@bot.command(name="removeuser")
@commands.has_any_role(ADMIN_ROLE_ID)
async def remove_user_cmd(ctx, member: discord.Member = None):
    if not member:
        await ctx.reply("❌ Gunakan: `!removeuser @user`"); return
    success = db_remove_user(str(member.id))
    if not success:
        await ctx.reply(f"❌ **{member.display_name}** tidak ada di database."); return
    embed = discord.Embed(title="🗑️ User Dihapus!", color=0xE74C3C)
    embed.set_author(name=member.display_name, icon_url=member.display_avatar.url)
    embed.add_field(name="👤 Player", value=member.mention,                        inline=True)
    embed.add_field(name="🗑️ Status", value="Dihapus dari leaderboard & database", inline=True)
    embed.set_footer(text=f"Dihapus oleh {ctx.author.display_name}")
    await ctx.reply(embed=embed)
    log_channel = discord.utils.get(ctx.guild.text_channels, name=LOG_CHANNEL_NAME)
    if log_channel:
        log_embed = discord.Embed(title="🗑️ Remove User — Log Admin", color=0xE74C3C, timestamp=datetime.now(timezone.utc))
        log_embed.add_field(name="👮 Admin",  value=ctx.author.mention,                 inline=True)
        log_embed.add_field(name="👤 Target", value=member.mention,                     inline=True)
        log_embed.add_field(name="🗑️ Aksi",   value="Dihapus dari database sepenuhnya", inline=False)
        log_embed.set_footer(text="SSRP Admin Log")
        await log_channel.send(embed=log_embed)

async def send_log(guild, action, admin, target, amount, before, after):
    log_channel = discord.utils.get(guild.text_channels, name=LOG_CHANNEL_NAME)
    if not log_channel: return
    color = 0x2ECC71 if action == "ADD" else 0xE74C3C
    icon  = "➕" if action == "ADD" else "➖"
    embed = discord.Embed(title=f"{icon} Poin {action} — Log Admin", color=color, timestamp=datetime.now(timezone.utc))
    embed.add_field(name="👮 Admin",   value=admin.mention,    inline=True)
    embed.add_field(name="👤 Target",  value=target.mention,   inline=True)
    embed.add_field(name="🔢 Jumlah",  value=f"{amount} poin", inline=True)
    embed.add_field(name="📉 Sebelum", value=f"{before} poin", inline=True)
    embed.add_field(name="📈 Sesudah", value=f"{after} poin",  inline=True)
    embed.set_footer(text="SSRP Admin Log")
    await log_channel.send(embed=embed)

@bot.command(name="addpoint")
@commands.has_any_role(ADMIN_ROLE_ID)
async def add_point_cmd(ctx, member: discord.Member = None, amount: int = 1):
    if not member:
        await ctx.reply("❌ Gunakan: `!addpoint @user [jumlah]`"); return
    if amount < 1:
        await ctx.reply("❌ Jumlah harus minimal 1."); return
    before_row = get_user_point(str(member.id))
    before     = before_row[0] if before_row else 0
    result     = manual_add_point(str(member.id), member.display_name, amount)
    after      = result[0]
    embed = discord.Embed(title="➕ Poin Ditambahkan!", color=0x2ECC71)
    embed.set_author(name=member.display_name, icon_url=member.display_avatar.url)
    embed.add_field(name="👤 Player",     value=member.mention,   inline=True)
    embed.add_field(name="➕ Ditambah",   value=f"{amount} poin", inline=True)
    embed.add_field(name="🏆 Total Baru", value=f"{after} poin",  inline=True)
    await ctx.reply(embed=embed)
    await send_log(ctx.guild, "ADD", ctx.author, member, amount, before, after)

@bot.command(name="removepoint")
@commands.has_any_role(ADMIN_ROLE_ID)
async def remove_point_cmd(ctx, member: discord.Member = None, amount: int = 1):
    if not member:
        await ctx.reply("❌ Gunakan: `!removepoint @user [jumlah]`"); return
    if amount < 1:
        await ctx.reply("❌ Jumlah harus minimal 1."); return
    row = get_user_point(str(member.id))
    if not row:
        await ctx.reply(f"❌ **{member.display_name}** belum punya data poin."); return
    before = row[0]
    pending_removals[str(ctx.author.id)] = {
        "target_id": str(member.id), "target_name": member.display_name,
        "target_obj": member, "amount": amount, "before": before,
        "expires": datetime.now(timezone.utc).timestamp() + 30
    }
    embed = discord.Embed(
        title="⚠️ Konfirmasi Hapus Poin",
        description=(
            f"Kamu akan menghapus **{amount} poin** dari {member.mention}\n"
            f"Poin sekarang: **{before}** → setelah: **{max(0, before - amount)}**\n\n"
            f"Ketik **`!confirm`** dalam **30 detik** untuk lanjutkan.\nKetik **`!cancel`** untuk batalkan."
        ),
        color=0xE67E22
    )
    await ctx.reply(embed=embed)

@bot.command(name="confirm")
@commands.has_any_role(ADMIN_ROLE_ID)
async def confirm_remove(ctx):
    admin_id = str(ctx.author.id)
    pending  = pending_removals.get(admin_id)
    if not pending:
        await ctx.reply("❌ Tidak ada pending removepoint untuk kamu."); return
    if datetime.now(timezone.utc).timestamp() > pending["expires"]:
        del pending_removals[admin_id]
        await ctx.reply("❌ Waktu konfirmasi habis (30 detik). Ulangi `!removepoint` lagi."); return
    result = manual_remove_point(pending["target_id"], pending["amount"])
    if not result:
        await ctx.reply("❌ Gagal, member tidak ditemukan di database."); return
    after  = result[0]; member = pending["target_obj"]
    before = pending["before"]; amount = pending["amount"]
    del pending_removals[admin_id]
    embed = discord.Embed(title="➖ Poin Dihapus!", color=0xE74C3C)
    embed.set_author(name=member.display_name, icon_url=member.display_avatar.url)
    embed.add_field(name="👤 Player",     value=member.mention,   inline=True)
    embed.add_field(name="➖ Dihapus",    value=f"{amount} poin", inline=True)
    embed.add_field(name="🏆 Total Baru", value=f"{after} poin",  inline=True)
    await ctx.reply(embed=embed)
    await send_log(ctx.guild, "REMOVE", ctx.author, member, amount, before, after)

@bot.command(name="cancel")
@commands.has_any_role(ADMIN_ROLE_ID)
async def cancel_remove(ctx):
    admin_id = str(ctx.author.id)
    if admin_id in pending_removals:
        del pending_removals[admin_id]
        await ctx.reply("✅ Removepoint dibatalkan.")
    else:
        await ctx.reply("❌ Tidak ada pending removepoint.")

@bot.command(name="ssrphelp")
async def ssrp_help(ctx):
    embed = discord.Embed(title="🤖 SSRP Checker Bot", color=0x5865F2)
    embed.add_field(name="📋 Cara Submit", value=(
        "Kirim screenshot di:\n"
        "• `#ssrp-approval`\n"
        "• `#outrider-ssrp-approval`\n"
        "Bot langsung kasih poin otomatis!"
    ), inline=False)
    embed.add_field(name="👤 Commands Member", value=(
        "`!point` — Cek poin + stats kamu\n"
        "`!point @user` — Cek poin member lain\n"
        "`!profile` — Kartu profil lengkap\n"
        "`!profile @user` — Profil member lain\n"
        "`!lb` — Leaderboard interaktif\n"
        "`!ssrphelp` — Bantuan ini"
    ), inline=False)
    embed.add_field(name="🔧 Commands Admin", value=(
        "`/admin` — Panel admin dengan tombol\n"
        "`!export excel/csv/all` — Export data\n"
        "`!inactive [hari]` — Cek tidak aktif\n"
        "`!resetweek` — Reset poin mingguan semua\n"
        "`!resetuser @user` — Reset poin 1 member → 0\n"
        "`!removeuser @user` — Hapus member dari LB\n"
        "`/announce` — Kirim pengumuman manual"
    ), inline=False)
    embed.add_field(name="🚨 Commands Darurat (Hanya saat bug!)", value=(
        "`!addpoint @user [n]` — Tambah poin manual\n"
        "`!removepoint @user [n]` — Kurangi poin manual\n"
        "`!confirm` — Konfirmasi removepoint\n"
        "`!cancel` — Batalkan removepoint\n"
        "⚠️ _Gunakan hanya jika ada kesalahan data!_"
    ), inline=False)
    embed.add_field(name="⚙️ Sistem", value=(
        f"• 1 submit = 1 poin (berapapun jumlah foto)\n"
        f"• Cooldown: **{SUBMIT_COOLDOWN} detik** antar submit\n"
        f"• Pengumuman otomatis setiap **Senin 00:00 WIB**\n"
        f"• Cek inaktif otomatis setiap Senin pagi"
    ), inline=False)
    await ctx.reply(embed=embed)

# ─── Run ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    bot.run(DISCORD_TOKEN)
